from pathlib import Path


def write_c_tail_decimal_to_e(
    input_filename: str = "label.xlsx",
    output_filename: str = "label_ten.xlsx",
) -> Path:
    """Copy label.xlsx and write decimal values from column C's last 6 hex chars to column E."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Please install openpyxl first: pip install openpyxl") from exc

    current_dir = Path(__file__).resolve().parent
    input_path = current_dir / input_filename
    output_path = current_dir / output_filename

    workbook = load_workbook(input_path)
    sheet = workbook.active

    for row in range(1, sheet.max_row + 1):
        raw_value = sheet.cell(row=row, column=3).value
        if raw_value is None:
            sheet.cell(row=row, column=5).value = None
            continue

        hex_text = str(raw_value).strip()
        tail = hex_text[-6:]

        try:
            decimal_value = int(tail, 16)
        except ValueError:
            sheet.cell(row=row, column=5).value = None
            continue

        sheet.cell(row=row, column=5).value = decimal_value

    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    print(write_c_tail_decimal_to_e())
